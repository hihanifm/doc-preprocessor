"""Helpers to inspect and filter arbitrary .xlsx sheets without loading huge distinct-value lists."""

from __future__ import annotations

import io
from typing import Any, Dict, List, Literal, Optional, Tuple

import openpyxl

Mode = Literal["contains", "equals", "not_contains", "starts_with"]


def normalize_cell(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def resolve_column_index(headers: List[str], column_name: str) -> int:
    for i, h in enumerate(headers):
        if h == column_name:
            return i
    target = column_name.lower().strip()
    for i, h in enumerate(headers):
        if h.lower().strip() == target:
            return i
    raise ValueError(f"Column not found: {column_name}")


def row_matches(cell_text: str, mode: Mode, needle: str) -> bool:
    if not needle.strip():
        return True
    c = cell_text.lower()
    n = needle.lower()
    if mode == "contains":
        return n in c
    if mode == "equals":
        return c == n
    if mode == "not_contains":
        return n not in c
    if mode == "starts_with":
        return c.startswith(n)
    return True


def filter_dict_rows(
    rows: List[Dict[str, Any]],
    column_key: str,
    mode: Mode,
    needle: str,
) -> List[Dict[str, Any]]:
    if not needle.strip():
        return list(rows)
    out: List[Dict[str, Any]] = []
    for row in rows:
        cell = normalize_cell(row.get(column_key, ""))
        if row_matches(cell, mode, needle):
            out.append(row)
    return out


def workbook_sheet_info(path: str, sheet_index: int = 0) -> Tuple[List[str], int, List[str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if sheet_index < 0 or sheet_index >= len(names):
            raise ValueError("Invalid sheet index")
        ws = wb[names[sheet_index]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], 0, names
        headers = [
            normalize_cell(h) if normalize_cell(h) else f"Column{i + 1}"
            for i, h in enumerate(header_row)
        ]
        count = 0
        for _ in rows_iter:
            count += 1
        return headers, count, names
    finally:
        wb.close()


def peek_distinct(
    path: str,
    column_name: str,
    sheet_index: int = 0,
    max_values: int = 30,
    max_scan_rows: int = 20000,
) -> Tuple[List[str], bool, int]:
    """
    Return up to max_values distinct non-empty cell strings from a column.
    Stops early when max_values reached or max_scan_rows scanned.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    distinct: List[str] = []
    seen: set[str] = set()
    scanned = 0
    truncated = False
    try:
        ws = wb[wb.sheetnames[sheet_index]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], False, 0
        headers = [
            normalize_cell(h) if normalize_cell(h) else f"Column{i + 1}"
            for i, h in enumerate(header_row)
        ]
        col_idx = resolve_column_index(headers, column_name)

        for row in rows_iter:
            scanned += 1
            if scanned > max_scan_rows:
                truncated = True
                break
            vals = list(row) + [None] * max(0, len(headers) - len(row))
            cell = normalize_cell(vals[col_idx] if col_idx < len(vals) else "")
            if not cell or cell in seen:
                continue
            seen.add(cell)
            distinct.append(cell)
            if len(distinct) >= max_values:
                truncated = True
                break
        return distinct, truncated, scanned
    finally:
        wb.close()


def filter_xlsx_to_bytes(
    path: str,
    column_name: str,
    mode: Mode,
    needle: str,
    sheet_index: int = 0,
) -> bytes:
    """Stream through a sheet and build a filtered workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[sheet_index]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return dict_rows_to_xlsx_bytes([], [])

        headers = [
            normalize_cell(h) if normalize_cell(h) else f"Column{i + 1}"
            for i, h in enumerate(header_row)
        ]
        col_idx = resolve_column_index(headers, column_name)

        matching: List[Dict[str, Any]] = []
        for row in rows_iter:
            vals = list(row)
            while len(vals) < len(headers):
                vals.append(None)
            cell = normalize_cell(vals[col_idx] if col_idx < len(vals) else "")
            if row_matches(cell, mode, needle if needle else ""):
                matching.append({headers[i]: vals[i] for i in range(len(headers))})

        return dict_rows_to_xlsx_bytes(matching, headers)
    finally:
        wb.close()


def sample_sheet_rows(
    path: str,
    sheet_index: int = 0,
    max_rows: int = 15,
) -> Tuple[List[str], List[List[Any]]]:
    """First header row + up to max_rows data rows for UI preview (read-only)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if sheet_index < 0 or sheet_index >= len(names):
            raise ValueError("Invalid sheet index")
        ws = wb[names[sheet_index]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], []
        headers = [
            normalize_cell(h) if normalize_cell(h) else f"Column{i + 1}"
            for i, h in enumerate(header_row)
        ]
        out: List[List[Any]] = []
        n = 0
        for row in rows_iter:
            if n >= max_rows:
                break
            vals = list(row)
            while len(vals) < len(headers):
                vals.append(None)
            out.append(vals[: len(headers)])
            n += 1
        return headers, out
    finally:
        wb.close()


def dict_rows_to_xlsx_bytes(
    rows: List[Dict[str, Any]],
    headers: List[str],
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered"

    if not headers:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    for r_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=col_idx, value=row.get(h, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def merge_xlsx_to_bytes(paths: List[str], filenames: List[str], add_source: bool = False) -> bytes:
    """Merge first sheet of each xlsx into one workbook. Raises ValueError if headers differ."""
    reference_headers: Optional[List[str]] = None
    all_rows: List[Dict] = []

    for path, fname in zip(paths, filenames):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        headers = [normalize_cell(h) for h in next(it, [])]
        if reference_headers is None:
            reference_headers = headers
        elif headers != reference_headers:
            raise ValueError(f"'{fname}' columns don't match the first file.")
        for row in it:
            d = {h: normalize_cell(v) for h, v in zip(reference_headers, row)}
            if add_source:
                d["Source file"] = fname
            all_rows.append(d)
        wb.close()

    if not reference_headers:
        raise ValueError("No data found in uploaded files.")

    out_headers = (["Source file"] + reference_headers) if add_source else reference_headers
    return dict_rows_to_xlsx_bytes(all_rows, out_headers)


def join_xlsx_to_bytes(target_path: str, source_path: str, key_col: str, columns_to_copy: List[str]) -> bytes:
    """LEFT JOIN target with source on key_col; copy selected columns. First source match wins."""
    src_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    src_ws = src_wb.worksheets[0]
    src_it = src_ws.iter_rows(values_only=True)
    src_headers = [normalize_cell(h) for h in next(src_it, [])]
    if key_col not in src_headers:
        raise ValueError(f"Key column '{key_col}' not found in source file.")
    src_key_idx = src_headers.index(key_col)
    lookup: Dict = {}
    for row in src_it:
        k = normalize_cell(row[src_key_idx])
        if k and k not in lookup:
            lookup[k] = {src_headers[i]: normalize_cell(v) for i, v in enumerate(row)}
    src_wb.close()

    tgt_wb = openpyxl.load_workbook(target_path, read_only=True, data_only=True)
    tgt_ws = tgt_wb.worksheets[0]
    tgt_it = tgt_ws.iter_rows(values_only=True)
    tgt_headers = [normalize_cell(h) for h in next(tgt_it, [])]
    if key_col not in tgt_headers:
        raise ValueError(f"Key column '{key_col}' not found in target file.")

    extra = [c for c in columns_to_copy if c not in tgt_headers]
    out_headers = tgt_headers + extra

    all_rows = []
    for row in tgt_it:
        d = {tgt_headers[i]: normalize_cell(v) for i, v in enumerate(row)}
        src_row = lookup.get(d.get(key_col, ""), {})
        for col in columns_to_copy:
            d[col] = src_row.get(col, "")
        all_rows.append(d)
    tgt_wb.close()

    return dict_rows_to_xlsx_bytes(all_rows, out_headers)
