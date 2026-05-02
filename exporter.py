import io
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


COLUMNS = [
    ("file_name", "File Name"),
    ("test_id", "Test ID"),
    ("test_name", "Test Name"),
    ("description", "Description"),
    ("preconditions", "Preconditions / Applicability"),
    ("procedure_steps", "Procedure / steps"),
    ("expected_results", "Expected results"),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_COLS = {"description", "preconditions", "procedure_steps", "expected_results"}


def to_excel(rows: List[dict]) -> bytes:
    """Convert a list of test case dicts to an Excel workbook and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Header row
    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 24

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if key in WRAP_COLS:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

    # Column widths
    col_widths = {
        "file_name": 30,
        "test_id": 20,
        "test_name": 35,
        "description": 50,
        "preconditions": 40,
        "procedure_steps": 50,
        "expected_results": 50,
    }
    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = col_widths.get(key, 20)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
